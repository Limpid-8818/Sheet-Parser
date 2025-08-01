import html
import os
from string import Template

import pandas as pd
from openpyxl import load_workbook
import csv
from utils import check_file_exists, check_file_format, get_default_title, determine_data_type


class SheetParser:
    """表格解析器，用于将多种格式的表格文件转换为HTML表格"""

    def __init__(self):
        """初始化解析器"""
        self.supported_formats = ['.xlsx', '.xls', '.csv', '.xlt', '.xltx', '.xlsb', '.xltm', '.xlam', '.et', '.ett',
                                  '.ets']
        self.html_template_path = 'templates/basic_table.html'
        with open(self.html_template_path, 'r', encoding='utf-8') as f:
            self.html_template = Template(f.read())

    def parse_file(self, file_path, output_file=None, title=None):
        """
        解析表格文件并生成HTML

        Args:
            file_path: 表格文件路径
            output_file: 输出HTML文件路径，默认为None，此时返回HTML字符串
            title: HTML页面标题，默认为文件名

        Returns:
            生成的HTML字符串(如果output_file为None)
        """
        check_file_exists(file_path)

        check_file_format(file_path, self.supported_formats)

        if title is None:
            title = get_default_title(file_path)

        # 根据文件格式选择解析方法
        if os.path.splitext(file_path)[1].lower() in ['.xlsx', '.xls', '.xlt', '.xltx', '.xlsb', '.xltm', '.xlam',
                                                      '.et', '.ett', '.ets']:
            sheets_data = self._parse_excel(file_path)
        elif os.path.splitext(file_path)[1].lower() == '.csv':
            sheets_data = self._parse_csv(file_path)

        # 生成HTML内容
        content = self._generate_html_content(sheets_data)
        html = self.html_template.substitute(title=title, content=content)

        if output_file:
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write(html)
            return f"HTML已保存到: {output_file}"
        else:
            return html

    def _parse_excel(self, file_path):
        """解析Excel文件(.xlsx, .xls)"""
        xls = pd.ExcelFile(file_path)
        sheet_names = xls.sheet_names

        sheets_data = []

        # 使用openpyxl获取合并单元格信息
        wb = load_workbook(file_path, data_only=True)
        wb_with_formula = load_workbook(file_path, data_only=False)

        for sheet_name in sheet_names:
            df = xls.parse(sheet_name, header=None)

            # 处理空数据框
            if df.empty:
                sheets_data.append({
                    'name': sheet_name,
                    'data': []
                })
                continue

            # 获取合并单元格信息
            ws = wb[sheet_name]
            ws_with_formula = wb_with_formula[sheet_name]
            merged_cells = []
            if hasattr(ws, 'merged_cells'):
                for merged_range in ws.merged_cells.ranges:
                    min_col, min_row, max_col, max_row = merged_range.bounds
                    merged_cells.append({
                        'min_row': min_row - 1,  # 转换为0-based索引
                        'min_col': min_col - 1,
                        'max_row': max_row - 1,
                        'max_col': max_col - 1
                    })

            # 创建一个二维数组来跟踪哪些单元格已被处理
            rows_count = df.shape[0]
            cols_count = df.shape[1] if rows_count > 0 else 0
            cell_processed = [[False for _ in range(cols_count)] for _ in range(rows_count)]

            # 辅助函数：标记单元格为已处理，避免越界
            def mark_cell_processed(row, col):
                if 0 <= row < rows_count and 0 <= col < cols_count:
                    cell_processed[row][col] = True

            # 处理所有行（包括表头）
            data = []
            for row_idx, row in enumerate(df.itertuples(index=False)):
                processed_row = []
                for col_idx, value in enumerate(row):
                    # 如果已经处理过或超出范围，跳过
                    if row_idx >= rows_count or col_idx >= cols_count or cell_processed[row_idx][col_idx]:
                        continue

                    # 处理NaN值
                    if pd.isna(value):
                        cell_value = ''
                        cell_type = 'string'
                    else:
                        # 确定数据类型
                        cell_type = determine_data_type(value)
                        # 格式化日期
                        if cell_type == 'date':
                            cell_value = value.strftime('%Y-%m-%d')
                        else:
                            cell_value = str(value)

                    # 检查是否为合并单元格
                    is_merged = False
                    rowspan = 1
                    colspan = 1
                    for merged_cell in merged_cells:
                        if (row_idx >= merged_cell['min_row'] and row_idx <= merged_cell['max_row'] and
                                col_idx >= merged_cell['min_col'] and col_idx <= merged_cell['max_col']):
                            is_merged = True
                            rowspan = merged_cell['max_row'] - merged_cell['min_row'] + 1
                            colspan = merged_cell['max_col'] - merged_cell['min_col'] + 1
                            # 标记整个合并区域为已处理
                            for r in range(merged_cell['min_row'], merged_cell['max_row'] + 1):
                                for c in range(merged_cell['min_col'], merged_cell['max_col'] + 1):
                                    mark_cell_processed(r, c)
                            break

                    # 检查是否有超链接
                    hyperlink = ws.cell(row=row_idx + 1, column=col_idx + 1).hyperlink
                    if hyperlink:
                        cell_value = f'<a href="{hyperlink.target}" target="_blank">{cell_value}</a>'

                    # 检查是否有注释
                    comment = ws.cell(row=row_idx + 1, column=col_idx + 1).comment
                    comment_text = (comment.text or '').strip() if comment else ''

                    # 只在有注释时才包 <span>
                    if comment_text:
                        # 转义引号避免 HTML 注入
                        safe_comment = html.escape(comment_text, quote=True)
                        cell_value = f'<span class="comment" data-comment="{safe_comment}">{cell_value}</span>'

                    # 检查公式
                    formula = self._get_formula(ws_with_formula, row_idx + 1, col_idx + 1)

                    # 获取单元格样式
                    cell_style = self._get_cell_style(ws.cell(row=row_idx + 1, column=col_idx + 1))

                    processed_row.append({
                        'value': cell_value,
                        'type': cell_type,
                        'rowspan': rowspan,
                        'colspan': colspan,
                        'is_merged': is_merged,
                        'style': cell_style,
                        'comment': comment_text,
                        'formula': formula
                    })

                data.append(processed_row)

            sheets_data.append({
                'name': sheet_name,
                'data': data
            })

        return sheets_data

    def _parse_csv(self, file_path):
        """解析CSV文件"""
        sheets_data = []

        # 读取CSV文件
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            rows = list(reader)

        if not rows:
            sheets_data.append({
                'name': os.path.basename(file_path),
                'header': [],
                'data': []
            })
            return sheets_data

        CSV_BORDER_STYLE = 'border: 1px solid #ddd;'

        # 处理表头
        header_row = rows[0]
        processed_header = []
        for col_idx, value in enumerate(header_row):
            cell_type = determine_data_type(value)
            processed_header.append({
                'value': value,
                'type': cell_type,
                'colspan': 1,
                'rowspan': 1,
                'is_merged': False,
                'style': CSV_BORDER_STYLE
            })

        # 处理数据行
        data = []
        for row_idx, row in enumerate(rows[1:], 1):
            processed_row = []
            for col_idx, value in enumerate(row):
                cell_type = determine_data_type(value)
                processed_row.append({
                    'value': value,
                    'type': cell_type,
                    'rowspan': 1,
                    'colspan': 1,
                    'is_merged': False,
                    'style': CSV_BORDER_STYLE
                })
            data.append(processed_row)

        # CSV文件只有一个工作表
        sheets_data.append({
            'name': os.path.basename(file_path),
            'header': processed_header,
            'data': data
        })

        return sheets_data

    def _generate_html_content(self, sheets_data):
        """生成HTML内容"""
        content = ''

        for sheet in sheets_data:
            sheet_name = sheet['name']
            header = sheet.get('header', [])
            sheet_data = sheet['data']

            # 生成表格HTML
            table_html = '<div class="sheet-container">'
            table_html += f'<h2 class="sheet-title">{sheet_name}</h2>'
            table_html += '<table class="sheet-table">'

            if header:
                # 生成表头，考虑到应避免与Excel文件表头行格式产生冲突，此处仅对csv文件生效
                table_html += '<thead><tr>'
                for cell in header:
                    cell_class = ' '.join([
                        'merged-cell' if cell['is_merged'] else '',
                        'numeric-cell' if cell['type'] == 'numeric' else '',
                        'date-cell' if cell['type'] == 'date' else '',
                        'boolean-cell' if cell['type'] == 'boolean' else ''
                    ]).strip()
                    table_html += (
                        f'<th class="{cell_class}" '
                        f'colspan="{cell["colspan"]}" rowspan="{cell["rowspan"]}" '
                        f'style="{cell["style"]}">{cell["value"]}</th>'
                    )
                table_html += '</tr></thead>'

            # 生成数据行
            table_html += '<tbody>'
            for row in sheet_data:
                table_html += '<tr>'
                for cell in row:
                    cell_class = ' '.join([
                        'merged-cell' if cell['is_merged'] else '',
                        'numeric-cell' if cell['type'] == 'numeric' else '',
                        'date-cell' if cell['type'] == 'date' else '',
                        'boolean-cell' if cell['type'] == 'boolean' else '',
                        'commented-cell' if cell.get('comment') else '',
                        'formula-cell' if cell.get('formula') else ''
                    ]).strip()
                    comment_attr = f' data-comment="{html.escape(cell.get("comment", ""), quote=True)}"' if cell.get(
                        'comment') else ''
                    formula_attr = f' data-formula="{html.escape(cell.get("formula", ""), quote=True)}"' if cell.get(
                        "formula") else ''
                    table_html += (
                        f'<td class="{cell_class}"{comment_attr}{formula_attr} '
                        f'colspan="{cell["colspan"]}" rowspan="{cell["rowspan"]}" '
                        f'style="{cell["style"]}">{cell["value"]}</td>'
                    )
                table_html += '</tr>'
            table_html += '</tbody>'
            table_html += '</table>'
            table_html += '</div>'

            content += table_html

        return content

    def _get_cell_style(self, cell):
        """获取单元格的样式信息并转换为HTML style属性"""
        style = []

        # 字体样式
        font = cell.font
        if font.bold:
            style.append('font-weight: bold')
        if font.italic:
            style.append('font-style: italic')
        if font.underline:
            style.append('text-decoration: underline')

        # 处理字体颜色
        if font.color and font.color.rgb:
            color_style = self._get_rgb_style_value(font.color.rgb)
            if color_style:
                style.append(f'color: {color_style}')

        if font.size:
            style.append(f'font-size: {font.size}pt')

        # 单元格背景色
        fill = cell.fill
        if fill.patternType and fill.patternType != 'none':
            bg_style = self._get_rgb_style_value(fill.fgColor.rgb)
            if bg_style:
                style.append(f'background-color: {bg_style}')

        # 边框样式
        border = cell.border
        for side in ('left', 'right', 'top', 'bottom'):
            css_border = self._get_side_style(getattr(border, side))
            if css_border:
                style.append(f'border-{side}: {css_border}')

        # 对齐方式
        alignment = cell.alignment
        if alignment.horizontal:
            style.append(f'text-align: {alignment.horizontal}')
        if alignment.vertical:
            style.append(f'vertical-align: {alignment.vertical}')

        return '; '.join(style)

    def _get_rgb_style_value(self, rgb):
        """将RGB对象或字符串转换为CSS可用的颜色值"""
        if rgb is None:
            return None
        if hasattr(rgb, 'rgb'):
            rgb = rgb.rgb  # RGB对象
        if isinstance(rgb, str) and len(rgb) == 8:
            # 去掉前两位透明度，取后6位
            return f"#{rgb[2:]}"
        return None

    def _get_side_style(self, side):
        """
            将 openpyxl.styles.Side 对象转换成 CSS border 片段。
            如果无边框，返回 None。
        """
        if not side or not side.style or side.style == 'none':
            return None

        style = self._get_css_style_border(side)
        color = self._get_rgb_style_value(side.color.rgb) or 'black'
        return f'{style} {color}'

    @staticmethod
    def _get_css_style_border(border):
        border_style_map = {
            'dashDot': 'dashed',
            'dashDotDot': 'dashed',
            'dashed': 'dashed',
            'dotted': 'dotted',
            'double': 'double',
            'hair': 'solid',
            'medium': 'solid',
            'mediumDashDot': 'dashed',
            'mediumDashDotDot': 'dashed',
            'mediumDashed': 'dashed',
            'slantDashDot': 'dashed',
            'thick': 'solid',
            'thin': 'solid'
        }
        border_width_map = {
            'thin': '1px',
            'medium': '2px',
            'thick': '3px',
            'hair': '0.5px',
            'double': '3px'
        }
        width_key = border.style if border.style in {'thin', 'medium', 'thick', 'hair', 'double'} else 'thin'
        return border_style_map.get(border.style, 'solid') + ' ' + border_width_map.get(width_key, '1px')

    @staticmethod
    def _get_formula(ws_formula, row, col):
        """返回公式字符串，若非公式则返回 None"""
        cell = ws_formula.cell(row=row, column=col)
        if cell.data_type == 'f':
            return cell.value
        return None
